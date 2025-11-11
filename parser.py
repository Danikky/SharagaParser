import openpyxl as xl # Обработка файла xlsx
import urllib # Работают вместе 
import requests # Работают вместе 
import timeit # Чекаем скорость работы
import json # Обработка формата данных
import os # Хранение данных

url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTOrEM-h36wb2LmpE3nF6f5tQMKVrBxWSZceBa7Jl5BZd5VaAt3GsHBLefgq9VV8RMDDb0FQEWNQol-/pub?output=xlsx"

def save_sheet(data: dict, filename: str):
    with open(f"jsons/{filename}.json", 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def read_sheet(filename: str):
    with open(f"jsons/{filename}.json", 'r', encoding='utf-8') as f:
        return json.load(f)
    
def update_tables(url=url):
    if url:
        try:        
            urllib.request.urlretrieve(url, "tables.xlsx")
        except:
            print("Ошибка скачивания таблицы")
    
# ДАЖЕ КОММЕНТЫ ДОБАВИЛ, АРТЕМ НЕ ТУПИ
def parsing(url=None, sheet=None, is_save=False): 
    """Функция для извлечения данных из таблицы с расписанием уолледжа подмосковье\n
    Keyword arguments:\n
    ulr: Имеет приоритет над sheet, скачивает таблицу и берёт активный sheet.\n
    sheet: Обрабатывает конкретный sheet. Заранее инициализировать в переменную и передать в функцию.\n
    is_save: Если 'True', сохранияет обработанную таблицу в формате '.json' в папку 'jsons'.\n
    Return: Список групп.
    """
    
    # Скачивание тиаблицы по url
    if url:
        try:
            urllib.request.urlretrieve(url, "tables.xlsx")

            # Инициализация xlsx таблицы
            wb = xl.load_workbook("tables.xlsx", read_only=True)
            sheet = wb.active
        except:
            print("Ошибка скачавания таблицы")
        print("Данные обновлены")
        
    groups = [] # Массим с группами
    
    # Инициализация групп
    row_id = 0 # номер строки
    
    move = 0
    if str(sheet["F1"]).lower == "понедельник":
        move = 5
    for row in sheet.iter_rows():
        
        row_id += 1 
        col_id = 0 # номер колонки
        
        for i in row:

            # Обнаружение группы
            col_id += 1
            if row_id == 3 + move*0 or row_id == 24 + move*1 or row_id == 45 + move*2 or row_id == 66 + move*3: # Строки с названиями групп
                if col_id > 2: # Скипаем "ВРЕМЯ" и "ПАРА"
                    if i.value: # Проверяем Null значение
                        
                        # Присваиваем курс
                        if row_id == 3 + move*0:
                            curse = 1
                        if row_id == 24 + move*1:
                            curse = 2
                        if row_id == 45 + move*2:
                            curse = 3
                        if row_id == 66 + move*3:
                            curse = 4
                        
                        # Записываем данные в массив
                        groups.append({
                            "id": col_id,
                            "name": i.value,
                            "curse": curse,
                            '1': None,
                            '2': None,
                            '3': None,
                            '4': None,
                            '5': None,
                            '6': None
                        })
    
    # Распределение пар и преподов по группам
    for group in groups:
        # Вычисление нужно строки 
        if group["curse"] == 1:
            row = 3
        if group["curse"] == 2:
            row = 24
        if group["curse"] == 3:
            row = 45
        if group["curse"] == 4:
            row = 66
            
        # Присваивание значений
        for i in range(5):
            group[str(i+1)] = {
                "para": sheet.cell(row=row + i*3 + 3, column=group["id"]).value,
                "teacher": sheet.cell(row=row + i*3 + 3, column=group["id"]).value
                }
    
    # Сохранение данных
    if is_save:
        save_sheet(data=groups, filename=sheet.title)
    
    # Возвращает списко групп студентов
    return groups    

def test_url():
    for i in parsing(url=url):
        print(f"""
| Группа: {i["name"]} | Курс: {i["curse"]} |
|1| {i['1']}
|2| {i['2']}
|3| {i['3']}
|4| {i['4']}
|5| {i['5']}
|6| {i['6']}
""")
    print("Время обработки таблицы: ", timeit.timeit(lambda: parsing(url), number=1))

def test_sheets():
    wb = xl.load_workbook("tables.xlsx", read_only=True)
    sheets = wb.sheetnames
    for sheet in sheets:
        print(f"// ТАБЛИЦА {sheet} //")
        for i in parsing(sheet=wb[sheet]):
            print(f"""
| Группа: {i["name"]} | Курс: {i["curse"]} |
|1| {i['1']}
|2| {i['2']}
|3| {i['3']}
|4| {i['4']}
|5| {i['5']}
|6| {i['6']}
        """)

def test_save():
    wb = xl.load_workbook("tables.xlsx", read_only=True)
    sheets = wb.sheetnames
    for sheet in sheets:
        print(f"// ТАБЛИЦА {sheet} ОБРАБАТЫВАЕТСЯ //")
        parsing(sheet=wb[sheet], is_save=True)
        print(f"// ТАБЛИЦА {sheet} СОХРАНЕНА //")
        print()
        
if __name__ == "__main__":
    # test_url()
    # test_sheets()
    test_save()
    # update_tables()
    pass